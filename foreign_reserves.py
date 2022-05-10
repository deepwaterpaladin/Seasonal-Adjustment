import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

# Canada's official international reserves, millions of United States dollars, Bank of Canada, daily (x 1,000,000)
# Date range A13:A1078
# B – Total, Canada's reserves
# C – Convertible foreign currencies, USD
# D - Convertible foreign currencies, other than United States
# E – Gold
# F –  Special draws
# G – Reserve position in the International Monetary Fund (IMF)

class ForeignReserves:
    def __init__(self):
        self.wb = load_workbook(filename='data/foreign_reserves.xlsx')
        self.ws = self.wb.active
        self.date_col = self.ws["A"]
        self.total_reserves = self.ws["B"]
        self.usd_reserves = self.ws["C"]
        self.non_usd_reserves = self.ws["D"]
        self.gold_reserves = self.ws["E"]
        self.special_draws = self.ws["F"]
        self.imf_reserves = self.ws["G"]
    
class TotalReserves(ForeignReserves):
    def __init__(self):
        super().__init__()
        self.name= self.ws["B10"]
        self.data = pd.DataFrame({"Date": [i.value for i in self.date_col[13:1078]], "Total": [i.value for i in self.total_reserves[13:1078]]})

    def plot(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["Total"], label="Total Foreign Reserves")
        self.ax.set_title(self.ws["A1"].value)
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Total Foreign Reserves (in millions of US Dollars)")
        self.ax.grid(True)
        self.ax.legend()
    
    def four_week_plot(self):
        self.four_week_data = pd.DataFrame({"Date": [i.value for i in self.date_col[1074:1078]], "Total": [i.value for i in self.total_reserves[1074:1078]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.four_week_data["Date"], self.four_week_data["Total"], label="Total Foreign Reserves")
        self.ax.set_title("Total Foreign Reserves over the last 4 weeks")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Total Foreign Reserves (in millions of US Dollars)")
        self.ax.grid(True)
        self.ax.legend()
    
    def three_month_plot(self):
        self.three_month_data = pd.DataFrame({"Date": [i.value for i in self.date_col[1066:1078]], "Total": [i.value for i in self.total_reserves[1066:1078]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_month_data["Date"], self.three_month_data["Total"], label="Total Foreign Reserves")
        self.ax.set_title("Total Foreign Reserves over the last 3 months")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Total Foreign Reserves (in millions of US Dollars)")
        self.ax.grid(True)
        self.ax.legend()
    
    def six_month_plot(self):
        self.six_month_data = pd.DataFrame({"Date": [i.value for i in self.date_col[1054:1078]], "Total": [i.value for i in self.total_reserves[1054:1078]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_data["Date"], self.six_month_data["Total"], label="Total Foreign Reserves")
        self.ax.set_title("Total Foreign Reserves over the last 6 months")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Total Foreign Reserves (in millions of US Dollars)")
        self.ax.grid(True)
        self.ax.legend()
    
    def twelve_month_plot(self):
        self.twelve_month_data = pd.DataFrame({"Date": [i.value for i in self.date_col[1030:1078]], "Total": [i.value for i in self.total_reserves[1030:1078]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.twelve_month_data["Date"], self.twelve_month_data["Total"], label="Total Foreign Reserves")
        self.ax.set_title("Total Foreign Reserves over the last 12 months")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Total Foreign Reserves (in millions of US Dollars)")
        self.ax.grid(True)
        self.ax.legend()
    
    def two_year_plot(self):
        self.two_year_data = pd.DataFrame({"Date": [i.value for i in self.date_col[977:1078]], "Total": [i.value for i in self.total_reserves[977:1078]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.two_year_data["Date"], self.two_year_data["Total"], label="Total Foreign Reserves")
        self.ax.set_title("Total Foreign Reserves over the last 2 years")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Total Foreign Reserves (in millions of US Dollars)")
        self.ax.grid(True)
        self.ax.legend()