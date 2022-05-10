import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

# Date range A13-A1346

class ExchangeRate:
    def __init__(self):
        self.wb = load_workbook(filename='data/exchange_rate.xlsx')
        self.ws = self.wb.active
        self.date_col = self.ws['A']
        pass

class AUD(ExchangeRate):
    '''
    Australian Dollar x Canadian Dollar, daily.
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['B10'].value
        self.AUD_col = self.ws['B']
        self.AUD_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.AUD_col[13:1346]]})

    def plot_AUD(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.AUD_db['Date'], self.AUD_db['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the AUD/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.AUD_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the AUD/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.AUD_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 AUD')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the AUD/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.AUD_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the AUD/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.AUD_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the AUD/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.AUD_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the AUD/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.AUD_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the AUD/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.AUD_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.AUD_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 AUD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('AUD/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

class YUAN(ExchangeRate):
    '''
    Chinese Renminbi x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['C10'].value
        self.YUAN_col = self.ws['C']
        self.YUAN_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.YUAN_col[13:1346]]})
    
    def plot_YUAN(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.YUAN_db['Date'], self.YUAN_db['Dollars'], label='1 Yuan')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the YUAN/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.YUAN_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 YUAN')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the YUAN/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.YUAN_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 YUAN')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the YUAN/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.YUAN_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 YUAN')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the YUAN/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.YUAN_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 YUAN')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the YUAN/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.YUAN_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 YUAN')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the YUAN/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.YUAN_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 YUAN')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the YUAN/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.YUAN_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 YUAN')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.YUAN_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 YUAN')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('YUAN/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

class INR(ExchangeRate):
    '''
    Indian Rupee x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['D10'].value
        self.INR_col = self.ws['D']
        self.INR_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.INR_col[13:1346]]})
    
    def plot_INR(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.INR_db['Date'], self.INR_db['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the INR/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.INR_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the INR/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.INR_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 INR')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the INR/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.INR_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the INR/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.INR_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the INR/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.INR_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the INR/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.INR_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the INR/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.INR_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.INR_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 INR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('INR/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

class RUB(ExchangeRate):
    '''
    Russian Ruble x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['E10'].value
        self.RUB_col = self.ws['E']
        self.RUB_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.RUB_col[13:1346]]})
    
    def plot_RUB(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.RUB_db['Date'], self.RUB_db['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the RUB/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.RUB_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the RUB/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.RUB_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 RUB')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the RUB/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.RUB_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the RUB/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.RUB_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the RUB/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.RUB_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the RUB/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.RUB_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the RUB/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.RUB_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.RUB_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 RUB')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('RUB/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

class ZAR(ExchangeRate):
    '''
    South African Rand x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['F10'].value
        self.ZAR_col = self.ws['F']
        self.ZAR_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.ZAR_col[13:1346]]})
    
    def plot_ZAR(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ZAR_db['Date'], self.ZAR_db['Dollars'], label='1 Rand')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the ZAR/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.ZAR_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 ZAR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the ZAR/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.ZAR_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 ZAR')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the ZAR/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.ZAR_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 ZAR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the ZAR/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.ZAR_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the ZAR/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.ZAR_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 ZAR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the ZAR/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.ZAR_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 ZAR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the ZAR/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.ZAR_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 ZAR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.ZAR_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 ZAR')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('ZAR/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

class KRW(ExchangeRate):
    '''
    South Korean Won x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['G10'].value
        self.KRW_col = self.ws['G']
        self.KRW_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.KRW_col[13:1346]]})
    
    def plot_KRW(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.KRW_db['Date'], self.KRW_db['Dollars'], label='1 Won')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the KRW/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.KRW_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 KRW')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the KRW/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.KRW_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 KRW')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the KRW/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.KRW_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 KRW')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the KRW/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.KRW_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 KRW')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the KRW/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.KRW_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 KRW')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the KRW/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.KRW_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 KRW')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the KRW/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.KRW_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 KRW')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.KRW_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 KRW')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('KRW/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
class TRY(ExchangeRate):
    '''
    Turkish Lira x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['H10'].value
        self.TRY_col = self.ws['H']
        self.TRY_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.TRY_col[13:1346]]})
    
    def plot_TRY(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.TRY_db['Date'], self.TRY_db['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the Lira/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.TRY_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the TRY/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.TRY_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 TRY')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('TRY/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the Lira/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.TRY_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the Lira/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.TRY_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the Lira/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.TRY_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the Lira/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.TRY_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the Lira/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.TRY_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.TRY_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 Lira')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('Lira/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
class GBP(ExchangeRate):
    '''
    British Pound x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['I10'].value
        self.GBP_col = self.ws['I']
        self.GBP_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.GBP_col[13:1346]]})
    
    def plot_GBP(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.GBP_db['Date'], self.GBP_db['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the GBP/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.GBP_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the GBP/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.GBP_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 GBP')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the GBP/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.GBP_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the GBP/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.GBP_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the GBP/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.GBP_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the GBP/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.GBP_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the GBP/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.GBP_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.GBP_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 GBP')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('GBP/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

class USD(ExchangeRate):
    '''
    United States Dollar x Canadian Dollar
    '''
    def __init__(self):
        super().__init__()
        self.name = self.ws['J10'].value
        self.USD_col = self.ws['J']
        self.USD_db = pd.DataFrame({'Date': [i.value for i in self.date_col[13:1346]], 'Dollars': [i.value for i in self.USD_col[13:1346]]})
    
    def plot_USD(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.USD_db['Date'], self.USD_db['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def seven_day_plot(self) -> None:
        '''
        Creates a plot of the 7 most recent days of the USD/CAD exchange rate.
        '''
        self.seven_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1339:1346]], 'Dollars': [i.value for i in self.USD_col[1339:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.seven_day_db['Date'], self.seven_day_db['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 7 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    def fourteen_day_plot(self) -> None:
        '''
        Creates a plot of the 14 most recent days of the USD/CAD exchange rate.
        '''
        self.fourteen_day_db = pd.DataFrame({'Date': [i.value for i in self.date_col[1332:1346]], 'Dollars': [i.value for i in self.USD_col[1332:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fourteen_day_db['Date'], self.fourteen_day_db['Dollars'], label='1 USD')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 14 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def thirty_day_plot(self):
        '''
        Creates a plot of the 30 most recent days of the USD/CAD exchange rate.
        '''
        self.thirty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1316:1346]], 'Dollars': [i.value for i in self.USD_col[1316:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.thirty_day_range['Date'], self.thirty_day_range['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 30 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def sixty_day_plot(self):
        '''
        Creates a plot of the 60 most recent days of the USD/CAD exchange rate.
        '''
        self.sixty_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1286:1346]], 'Dollars': [i.value for i in self.USD_col[1286:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.sixty_day_range['Date'], self.sixty_day_range['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 60 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def ninety_day_plot(self):
        '''
        Creates a plot of the 90 most recent days of the USD/CAD exchange rate.
        '''
        self.ninety_day_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1256:1346]], 'Dollars': [i.value for i in self.USD_col[1256:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.ninety_day_range['Date'], self.ninety_day_range['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 90 Day Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
        
    def six_month_plot(self):
        '''
        Creates a plot of the 6 most recent months of the USD/CAD exchange rate.
        '''
        self.six_month_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1226:1346]], 'Dollars': [i.value for i in self.USD_col[1226:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.six_month_range['Date'], self.six_month_range['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 6 Month Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()
    
    def one_year_plot(self):
        '''
        Creates a plot of the most recent year of the USD/CAD exchange rate.
        '''
        self.one_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[1086:1346]], 'Dollars': [i.value for i in self.USD_col[1086:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.one_year_range['Date'], self.one_year_range['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 1 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()

    def three_year_plot(self):
        self.three_year_range = pd.DataFrame({'Date': [i.value for i in self.date_col[566:1346]], 'Dollars': [i.value for i in self.USD_col[566:1346]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.three_year_range['Date'], self.three_year_range['Dollars'], label='1 USD')
        self.ax.set_title(self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('CAD Dollars')
        self.ax.set_title('USD/CAD 3 Year Exchange Rate')
        self.ax.grid(True)
        self.ax.legend()