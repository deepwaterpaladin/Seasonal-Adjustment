import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook


class Commodity:
    def __init__(self):
        self.wb = load_workbook(filename='data/prices.xlsx') 
        self.ws = self.wb.active
        self.date_col = self.ws["B"] 
        pass

class Energy(Commodity):
    def __init__(self):
        super().__init__()
        self.name = self.ws["E9"].value
        self.energy_col = self.ws["E"]
        pass
    def plot_commodity(self):
        self.energy_db = pd.DataFrame({'Date': [i.value for i in self.date_col[11:279]],'Price': [i.value for i in self.energy_col[11:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.energy_db["Date"], self.energy_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices')
        self.ax.legend()
    def plot_twelve_months(self):
        self.energy_db = pd.DataFrame({'Date': [i.value for i in self.date_col[267:279]],'Price': [i.value for i in self.energy_col[267:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.energy_db["Date"], self.energy_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 12 Months')
        self.ax.legend()
    def plot_twenty_four_months(self):
        self.energy_db = pd.DataFrame({'Date': [i.value for i in self.date_col[255:279]],'Price': [i.value for i in self.energy_col[255:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.energy_db["Date"], self.energy_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 24 Months')
        self.ax.legend()
    def plot_five_years(self):
        self.energy_db = pd.DataFrame({'Date': [i.value for i in self.date_col[219:279]],'Price': [i.value for i in self.energy_col[219:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.energy_db["Date"], self.energy_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 5 Years')
        self.ax.legend()
    def plot_ten_years(self):
        self.energy_db = pd.DataFrame({'Date': [i.value for i in self.date_col[159:279]],'Price': [i.value for i in self.energy_col[159:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.energy_db["Date"], self.energy_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 10 Years')
        self.ax.legend()

class Metals(Commodity):
    def __init__(self):
        super().__init__()
        self.name = self.ws["F9"].value
        self.metals_col = self.ws["F"]
        pass
    def plot_commodity(self):
        self.metals_db = pd.DataFrame({'Date': [i.value for i in self.date_col[11:279]],'Price': [i.value for i in self.metals_col[11:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.metals_db["Date"], self.metals_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices')
        self.ax.legend()
    def plot_twelve_months(self):
        self.metals_db = pd.DataFrame({'Date': [i.value for i in self.date_col[267:279]],'Price': [i.value for i in self.metals_col[267:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.metals_db["Date"], self.metals_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 12 Months')
        self.ax.legend()
    def plot_twenty_four_months(self):
        self.metals_db = pd.DataFrame({'Date': [i.value for i in self.date_col[255:279]],'Price': [i.value for i in self.metals_col[255:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.metals_db["Date"], self.metals_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 24 Months')
        self.ax.legend()
    def plot_five_years(self):
        self.metals_db = pd.DataFrame({'Date': [i.value for i in self.date_col[219:279]],'Price': [i.value for i in self.metals_col[219:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.metals_db["Date"], self.metals_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 5 Years')
        self.ax.legend()
    def plot_ten_years(self):
        self.metals_db = pd.DataFrame({'Date': [i.value for i in self.date_col[159:279]],'Price': [i.value for i in self.metals_col[159:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.metals_db["Date"], self.metals_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 10 Years')
        self.ax.legend()

class Agricultural(Commodity):
    def __init__(self):
        super().__init__()
        self.name = self.ws["G9"].value
        self.agricultural_col = self.ws["G"]
        pass
    def plot_commodity(self):
        self.argo_db = pd.DataFrame({'Date': [i.value for i in self.date_col[11:279]],'Price': [i.value for i in self.agricultural_col[11:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.argo_db["Date"], self.argo_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, 2000 - 2022')
        self.ax.legend()
    def plot_twelve_months(self):
        self.argo_db = pd.DataFrame({'Date': [i.value for i in self.date_col[267:279]],'Price': [i.value for i in self.agricultural_col[267:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.argo_db["Date"], self.argo_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 12 Months')
        self.ax.legend()
    def plot_twenty_four_months(self):
        self.argo_db = pd.DataFrame({'Date': [i.value for i in self.date_col[255:279]],'Price': [i.value for i in self.agricultural_col[255:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.argo_db["Date"], self.argo_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 24 Months')
        self.ax.legend()
    def plot_five_years(self):
        self.argo_db = pd.DataFrame({'Date': [i.value for i in self.date_col[219:279]],'Price': [i.value for i in self.agricultural_col[219:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.argo_db["Date"], self.argo_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 5 Years')
        self.ax.legend()
    def plot_ten_years(self):
        self.argo_db = pd.DataFrame({'Date': [i.value for i in self.date_col[159:279]],'Price': [i.value for i in self.agricultural_col[159:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.argo_db["Date"], self.argo_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 10 Years')
        self.ax.legend()
    
class Fish(Commodity):
    def __init__(self):
        super().__init__()
        self.name = self.ws["H9"].value
        self.fish_col = self.ws["H"]
        pass
    def plot_commodity(self):
        self.fish_db = pd.DataFrame({'Date': [i.value for i in self.date_col[11:279]],'Price': [i.value for i in self.fish_col[11:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fish_db["Date"], self.fish_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, 2000 - 2022')
        self.ax.legend()  
    def plot_twelve_months(self):
        self.fish_db = pd.DataFrame({'Date': [i.value for i in self.date_col[267:279]],'Price': [i.value for i in self.fish_col[267:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fish_db["Date"], self.fish_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 12 Months')
        self.ax.legend()  
    def plot_twenty_four_months(self):
        self.fish_db = pd.DataFrame({'Date': [i.value for i in self.date_col[255:279]],'Price': [i.value for i in self.fish_col[255:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fish_db["Date"], self.fish_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 24 Months')
        self.ax.legend()  
    def plot_five_years(self):
        self.fish_db = pd.DataFrame({'Date': [i.value for i in self.date_col[219:279]],'Price': [i.value for i in self.fish_col[219:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fish_db["Date"], self.fish_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 5 Years')
        self.ax.legend() 
    def plot_ten_years(self):
        self.fish_db = pd.DataFrame({'Date': [i.value for i in self.date_col[159:279]],'Price': [i.value for i in self.fish_col[159:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.fish_db["Date"], self.fish_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 10 Years')
        self.ax.legend() 

class Forestry(Commodity):
    def __init__(self):
        super().__init__()
        self.name = self.ws["I9"].value
        self.forestry_col = self.ws["I"]
    def plot_commodity(self):
        self.forest_db = pd.DataFrame({'Date': [i.value for i in self.date_col[11:279]],'Price': [i.value for i in self.forestry_col[11:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.forest_db["Date"], self.forest_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, 2000 - 2022')
        self.ax.legend()
    def plot_twelve_months(self):
        self.forest_db = pd.DataFrame({'Date': [i.value for i in self.date_col[267:279]],'Price': [i.value for i in self.forestry_col[267:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.forest_db["Date"], self.forest_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 12 Months')
        self.ax.legend()
    def plot_twenty_four_months(self):
        self.forest_db = pd.DataFrame({'Date': [i.value for i in self.date_col[255:279]],'Price': [i.value for i in self.forestry_col[255:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.forest_db["Date"], self.forest_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 24 Months')
        self.ax.legend()
    def plot_five_years(self):
        self.forest_db = pd.DataFrame({'Date': [i.value for i in self.date_col[219:279]],'Price': [i.value for i in self.forestry_col[219:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.forest_db["Date"], self.forest_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 5 Years')
        self.ax.legend()
    def plot_ten_years(self):
        self.forest_db = pd.DataFrame({'Date': [i.value for i in self.date_col[159:279]],'Price': [i.value for i in self.forestry_col[159:279]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.forest_db["Date"], self.forest_db["Price"], label=self.name)
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Price (Index: 1972 = $100)')
        self.ax.set_title(f'{self.name} Prices, Previous 10 Years')
        self.ax.legend()

class CommodityPrinter(Commodity):
    def __init__(self):
        super().__init__() 
        self.energy = Energy()
        self.metals = Metals()
        self.agro = Agricultural()
        self.fish = Fish()
        self.forest = Forestry()
    def print_energy(self):
        self.energy.plot_commodity()
        self.energy.plot_twelve_months()
        self.energy.plot_twenty_four_months()
        self.energy.plot_five_years()
        self.energy.plot_ten_years()
    def print_metals(self):
        self.metals.plot_commodity()
        self.metals.plot_twelve_months()
        self.metals.plot_twenty_four_months()
        self.metals.plot_five_years()
        self.metals.plot_ten_years()
    def print_agro(self):
        self.agro.plot_commodity()
        self.agro.plot_twelve_months()
        self.agro.plot_twenty_four_months()
        self.agro.plot_five_years()
        self.agro.plot_ten_years()
    def print_fish(self):
        self.fish.plot_commodity()
        self.fish.plot_twelve_months()
        self.fish.plot_twenty_four_months()
        self.fish.plot_five_years()
        self.fish.plot_ten_years()
    def print_forest(self):
        self.forest.plot_commodity()
        self.forest.plot_twelve_months()
        self.forest.plot_twenty_four_months()
        self.forest.plot_five_years()
        self.forest.plot_ten_years()
