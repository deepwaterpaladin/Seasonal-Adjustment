from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
wb = load_workbook(filename='prices.xlsx')
ws = wb.active
def get_energy_data(ws):
    date_col = ws["B"]
    commodities = ws["E"]
    df = pd.DataFrame({'Date': [i.value for i in date_col[11:237]], 'Price': [i.value for i in commodities[11:237]]})
    return df
#commodities = ws["C"]
#date_col = ws["B"]
#df = pd.DataFrame({'Date': [i.value for i in date_col[11:237]], 'Price': [i.value for i in commodities[11:237]]})
#meme = {"Commodity data": df}
print(get_energy_data(ws))

