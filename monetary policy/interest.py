import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

# 05-11-2012 to 05-10-2022
# Date range A17:A2513
# Overnight Target Rate Column B
# Overnight Money Market Financing Rate Column C
# BOC Rate Column D
# Canadian Overnight Repo Rate Average (CORRA) (%) Column E
# Operating Band, Low Column F
# Operating Band, High Column G

class BOC:
    def __init__(self):
        self.wb = load_workbook(filename='data/rates.xlsx')
        self.ws = self.wb.active
        self.date_range = self.ws['A'][17:2513]

class OperatingBand(BOC):
    def __init__(self):
        super().__init__()
        self.data = pd.DataFrame({"Date": [i.value for i in self.ws['A'][17:2513]], "Low": [i.value for i in self.ws["F"][17:2513]], "High": [i.value for i in self.ws["G"][17:2513]]})
    
    def plot(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data['Date'], self.data['High'], label='High')
        self.ax.plot(self.data['Date'], self.data['Low'], label='Low')
        self.ax.set_title('Bank of Canada Operating Band')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Interest Rate %')
        self.ax.grid()
        self.ax.legend()

class OvernightRate(BOC):
    def __init__(self):
        super().__init__()
        self.data = pd.DataFrame({"Date": [i.value for i in self.ws['A'][17:2513]], "Overnight Rate, Target": [i.value for i in self.ws["B"][17:2513]], "Overnight Rate, Actual": [i.value for i in self.ws["C"][17:2513]]})
    
    def plot(self):
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data['Date'], self.data['Overnight Rate, Target'], label='Target')
        self.ax.plot(self.data['Date'], self.data['Overnight Rate, Actual'], label='Actual')
        self.ax.set_title('Bank of Canada Overnight Rate')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Interest Rate %')
        self.ax.grid()
        self.ax.legend()

    def plot_with_corra(self):
        self.corra = pd.DataFrame({"CORRA": [i.value for i in self.ws["E"][17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data['Date'], self.data['Overnight Rate, Target'], label='Target')
        self.ax.plot(self.data['Date'], self.data['Overnight Rate, Actual'], label='Actual')
        self.ax.plot(self.data['Date'], self.corra['CORRA'], label='CORRA')
        self.ax.set_title('Bank of Canada Overnight Rate')
        self.ax.set_xlabel('Date')
        self.ax.set_ylabel('Interest Rate %')
        self.ax.grid()
        self.ax.legend()