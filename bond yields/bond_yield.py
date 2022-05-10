from operator import index
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

# 05-10-2012 to 05-10-2022
# Date range A17:A2513
# 2-Year Yield Column B
# 3-Year Yield Column C
# 5-Year Yield Column D
# 7-Year Yield Column E
# 10-Year Yield Column F
# 30-Year Yield Column G

class Bond:
    def __init__(self):
        self.wb = load_workbook(filename='data/yield.xlsx')
        self.ws = self.wb.active
        self.date_col = self.ws["A"]
    
    def plot_all(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.ws["A"][17:2513]], "2-Year": [i.value for i in self.ws["B"][17:2513]], "3-Year": [i.value for i in self.ws["C"][17:2513]], "5-Year": [i.value for i in self.ws["D"][17:2513]], "7-Year": [i.value for i in self.ws["E"][17:2513]], "10-Year": [i.value for i in self.ws["F"][17:2513]], "30-Year": [i.value for i in self.ws["G"][17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["2-Year"], label="2-Year")
        self.ax.plot(self.data["Date"], self.data["3-Year"], label="3-Year")
        self.ax.plot(self.data["Date"], self.data["5-Year"], label="5-Year")
        self.ax.plot(self.data["Date"], self.data["7-Year"], label="7-Year")
        self.ax.plot(self.data["Date"], self.data["10-Year"], label="10-Year")
        self.ax.plot(self.data["Date"], self.data["30-Year"], label="30-Year")
        self.ax.set_title("Bond Yields")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield")
        self.ax.grid(True)
        self.ax.legend()
    
    def plot_all_one_week(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.ws["A"][2508:2513]], "2-Year": [i.value for i in self.ws["B"][2508:2513]], "3-Year": [i.value for i in self.ws["C"][2508:2513]], "5-Year": [i.value for i in self.ws["D"][2508:2513]], "7-Year": [i.value for i in self.ws["E"][2508:2513]], "10-Year": [i.value for i in self.ws["F"][2508:2513]], "30-Year": [i.value for i in self.ws["G"][2508:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["2-Year"], label="2-Year")
        self.ax.plot(self.data["Date"], self.data["3-Year"], label="3-Year")
        self.ax.plot(self.data["Date"], self.data["5-Year"], label="5-Year")
        self.ax.plot(self.data["Date"], self.data["7-Year"], label="7-Year")
        self.ax.plot(self.data["Date"], self.data["10-Year"], label="10-Year")
        self.ax.plot(self.data["Date"], self.data["30-Year"], label="30-Year")
        self.ax.set_title("Bond Yields")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield")
        self.ax.grid(True)
        self.ax.legend()

    def plot_all_twelve_months(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.ws["A"][2260:2513]], "2-Year": [i.value for i in self.ws["B"][2260:2513]], "3-Year": [i.value for i in self.ws["C"][2260:2513]], "5-Year": [i.value for i in self.ws["D"][2260:2513]], "7-Year": [i.value for i in self.ws["E"][2260:2513]], "10-Year": [i.value for i in self.ws["F"][2260:2513]], "30-Year": [i.value for i in self.ws["G"][2260:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["2-Year"], label="2-Year")
        self.ax.plot(self.data["Date"], self.data["3-Year"], label="3-Year")
        self.ax.plot(self.data["Date"], self.data["5-Year"], label="5-Year")
        self.ax.plot(self.data["Date"], self.data["7-Year"], label="7-Year")
        self.ax.plot(self.data["Date"], self.data["10-Year"], label="10-Year")
        self.ax.plot(self.data["Date"], self.data["30-Year"], label="30-Year")
        self.ax.set_title("Bond Yields")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield")
        self.ax.grid(True)
        self.ax.legend()

        

class ThirtyYear(Bond):
    def __init__(self):
        super().__init__()
        self.name= "30-year Bond Yield, daily"
        self.yield_col = self.ws["G"]
    
    def plot(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.date_col[17:2513]], "Yield": [i.value for i in self.yield_col[17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["Yield"], label="30-year Bond Yield, daily")
        self.ax.set_title("30-year Yield, May 2012 - May 2022")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield %")
        self.ax.grid(True)
        self.ax.legend()
    
    def plot2(self, dataframe):
        '''
        Takes a panda dataframe as an argument and plots the dataframe
        '''
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(dataframe["Date"], dataframe["Yield"], label="30-year Bond Yield, daily")
        self.ax.set_title("30-year Yield, May 2012 - May 2022")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield %")
        self.ax.grid(True)
        self.ax.legend()

class TenYear(Bond):
    def __init__(self):
        super().__init__()
        self.name= "10-year Bond Yield, daily"
        self.yield_col = self.ws["F"]
    def plot(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.date_col[17:2513]], "Yield": [i.value for i in self.yield_col[17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["Yield"], label="10-year Bond Yield, daily")
        self.ax.set_title("10-year Yield, May 2012 - May 2022")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield %")
        self.ax.grid(True)
        self.ax.legend()
        
class SevenYear(Bond):
    def __init__(self):
        super().__init__()
        self.name= "7-year Bond Yield, daily"
        self.yield_col = self.ws["E"]
    def plot(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.date_col[17:2513]], "Yield": [i.value for i in self.yield_col[17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["Yield"], label="7-year Bond Yield, daily")
        self.ax.set_title("7-year Yield, May 2012 - May 2022")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield %")
        self.ax.grid(True)
        self.ax.legend()

class FiveYear(Bond):
    def __init__(self):
        super().__init__()
        self.name= "5-year Bond Yield, daily"
        self.yield_col = self.ws["D"]
    
    def plot(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.date_col[17:2513]], "Yield": [i.value for i in self.yield_col[17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["Yield"], label="5-year Bond Yield, daily")
        self.ax.set_title("5-year Yield, May 2012 - May 2022")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield %")
        self.ax.grid(True)
        self.ax.legend()

class ThreeYear(Bond):
    def __init__(self):
        super().__init__()
        self.name= "3-year Bond Yield, daily"
        self.yield_col = self.ws["C"]

    def plot(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.date_col[17:2513]], "Yield": [i.value for i in self.yield_col[17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["Yield"], label="3-year Bond Yield, daily")
        self.ax.set_title("3-year Yield, May 2012 - May 2022")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield %")
        self.ax.grid(True)
        self.ax.legend()

class TwoYear(Bond):
    def __init__(self):
        super().__init__()
        self.name= "2-year Bond Yield, daily"
        self.yield_col = self.ws["B"]
    
    def plot(self):
        self.data = pd.DataFrame({"Date": [i.value for i in self.date_col[17:2513]], "Yield": [i.value for i in self.yield_col[17:2513]]})
        self.fig, self.ax = plt.subplots(figsize=(18,9))
        self.ax.plot(self.data["Date"], self.data["Yield"], label="2-year Bond Yield, daily")
        self.ax.set_title("2-year Yield, May 2012 - May 2022")
        self.ax.set_xlabel("Date")
        self.ax.set_ylabel("Yield %")
        self.ax.grid(True)
        self.ax.legend()


