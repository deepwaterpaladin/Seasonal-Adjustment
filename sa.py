from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
wb = load_workbook(filename='prices.xlsx')
ws = wb.active

def get_data_types(ws):
    headings = ws[9]
    date_col = ws["B"]
    data_dict = {}
    i = 1
    for i in range(1, len(headings)):
        data_dict.update({headings[i].coordinate[:1]: headings[i].value})
    full_dict = {}
    for i in data_dict.keys():
        date_coll = ws[i]
        full_dict.update({i: [j for j in date_coll[12:237]]})
    return full_dict

def get_date_range(ws):
    date_col = ws["B"]
    date_range = [i for i in date_col[12:237]]
    date_value = [i.value for i in date_col[12:237]]
    return date_range, date_value
def test(ws):
    headings = ws[9]
    date_col = ws["B"]
    data_dict = {}
    i = 1
    for i in range(1, len(headings)):
        data_dict.update({headings[i].coordinate[:1]: headings[i].value})
    return data_dict

#print(get_data_types(ws))

def get_single_commodity(ws):
    column = input("Enter the column (as an uppercase letter): ")
    date_col = ws["B"]
    headings = ws[9]
    data_dict = {}
    target_range = ws[column][11:237]
    target_value = [i.value for i in target_range]
    for i in range(len(target_value)):
        data_dict.update({date_col[i+11].value: target_value[i]})
    return data_dict

print(get_single_commodity(ws))